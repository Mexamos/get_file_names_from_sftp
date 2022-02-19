from stat import S_ISDIR, S_ISREG

import paramiko
from openpyxl import Workbook

paramiko.util.log_to_file("paramiko.log")

host, port = "host", 22
transport = paramiko.Transport((host,port))
username, password = "username", "password"
transport.connect(None,username,password)
sftp = paramiko.SFTPClient.from_transport(transport)

wb = Workbook()
ws = wb.active
ws['A1'] = 'Названия файлов'
ws.column_dimensions['A'].width = 80

def get_all_files(sftp):
    folders = []
    while True:
        path = folders.pop() if len(folders) > 0 else '.'
        print(path)
        for entry in sftp.listdir_iter(path):
            mode = entry.st_mode
            if S_ISDIR(mode):
                start_path = path if path != '.' else ''
                next_path = f'{start_path}{entry.filename}/'
                folders.append(next_path)
            elif S_ISREG(mode):
                yield entry.filename
        
        if len(folders) == 0:
            break

count = 0
page = 2
for file_name in get_all_files(sftp) or []:
    ws[f'A{page}'] = file_name
    page += 1

    count += 1
    if not count % 500:
        print('count', count)
        wb.save('files_names.xlsx')

if sftp: sftp.close()
if transport: transport.close()
